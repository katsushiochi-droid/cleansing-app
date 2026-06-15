# -*- coding: utf-8 -*-
"""顧客データ クレンジング処理(UI 非依存)。

提供関数:
  read_input(file_obj, filename) -> pd.DataFrame   CSV/Excel を全列文字列で読む
  cleanse(df)                    -> CleanseResult    正規化・品質フラグ・集計
  build_excel(result)            -> io.BytesIO        書式付き4シートExcel
"""
import io
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass

import pandas as pd

EXPECTED_COLUMNS = [
    "customer_id", "name", "name_kana", "email", "phone", "postal_code",
    "address", "gender", "age", "annual_income", "registered_date",
    "last_purchase_date", "status", "notes",
]


def _to_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def read_input(file_obj, filename=None) -> pd.DataFrame:
    name = (filename or getattr(file_obj, "name", "") or "").lower()
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(file_obj, dtype=object)
    elif name.endswith(".csv"):
        raw = file_obj.read()
        if isinstance(raw, str):
            raw = raw.encode("utf-8")
        text = None
        for enc in ("utf-8-sig", "cp932"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("文字コードを判定できませんでした(UTF-8 か Shift_JIS で保存してください)")
        df = pd.read_csv(io.StringIO(text), dtype=object, keep_default_na=False)
    else:
        raise ValueError("CSV または Excel(.xlsx)ファイルを指定してください")
    df = df.fillna("")
    df = df.map(_to_str)
    df = df.astype(object)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@dataclass
class CleanseResult:
    raw_df: pd.DataFrame
    clean_df: pd.DataFrame
    review_df: pd.DataFrame
    report_df: pd.DataFrame
    summary_df: pd.DataFrame
    missing_columns: list


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
GENDER_MAP = {
    "男性": "男性", "男": "男性", "male": "男性", "m": "男性",
    "女性": "女性", "女": "女性", "female": "女性", "f": "女性",
    "その他": "その他", "未回答": "未回答", "": "未回答",
}
STATUS_MAP = {
    "active": "active", "有効": "active",
    "inactive": "inactive", "休眠": "inactive",
    "pending": "pending", "保留": "pending",
}
SENTINELS = {"", "not_a_date", "未購入", "9999-99-99", "0000-00-00", "null", "N/A"}


def _strip_all(s):
    if s is None:
        return ""
    return re.sub(r"^[\s　]+|[\s　]+$", "", str(s))


def cleanse(df: pd.DataFrame) -> CleanseResult:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.reset_index(drop=True)
    raw = df.copy()
    missing_columns = [c for c in EXPECTED_COLUMNS if c not in df.columns]

    n = len(df)
    flags = [[] for _ in range(n)]

    def flag(i, msg):
        flags[i].append(msg)

    for col in df.columns:
        df[col] = df[col].map(_strip_all)

    if "name" in df.columns:
        df["name"] = df["name"].str.replace(r"[\s　]+", " ", regex=True)

    if "name_kana" in df.columns:
        for i, v in enumerate(df["name_kana"]):
            if v.startswith("サンプル"):
                flag(i, "カナがプレースホルダ")

    if "email" in df.columns:
        df["email"] = df["email"].map(lambda s: s.replace("＠", "@").strip())
        for i, v in enumerate(df["email"]):
            if v == "":
                flag(i, "メール空欄")
            elif not EMAIL_RE.match(v):
                flag(i, "メール形式不正")

    if "phone" in df.columns:
        def clean_phone(s, i):
            if s == "":
                flag(i, "電話空欄")
                return ""
            t = s.replace("TEL:", "").replace("tel:", "")
            t = unicodedata.normalize("NFKC", t)
            t = t.replace("+81-", "0").replace("+81", "0")
            digits = re.sub(r"\D", "", t)
            if len(digits) == 11 and digits.startswith("0"):
                return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
            if len(digits) == 10 and digits.startswith("0"):
                return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
            flag(i, f"電話桁数不正({len(digits)}桁)")
            return digits
        df["phone"] = [clean_phone(v, i) for i, v in enumerate(df["phone"])]

    if "postal_code" in df.columns:
        def clean_postal(s, i):
            if s == "":
                flag(i, "郵便番号空欄")
                return ""
            t = unicodedata.normalize("NFKC", s)
            t = t.replace("〒", "").replace("ー", "-").replace("－", "-")
            digits = re.sub(r"\D", "", t)
            if len(digits) == 7:
                return f"{digits[:3]}-{digits[3:]}"
            flag(i, f"郵便番号桁数不正({len(digits)}桁)")
            return t
        df["postal_code"] = [clean_postal(v, i) for i, v in enumerate(df["postal_code"])]

    if "address" in df.columns:
        df["address"] = df["address"].map(lambda s: unicodedata.normalize("NFKC", s))

    if "gender" in df.columns:
        def clean_gender(s, i):
            key = s.strip().lower()
            if key in GENDER_MAP:
                return GENDER_MAP[key]
            flag(i, f"性別不明値({s})")
            return "未回答"
        df["gender"] = [clean_gender(v, i) for i, v in enumerate(df["gender"])]

    if "age" in df.columns:
        def clean_age(s, i):
            m = re.search(r"-?\d+", s)
            if not m:
                flag(i, "年齢空欄" if s == "" else f"年齢非数値({s})")
                return None
            num = int(m.group())
            if 1 <= num <= 120:
                return num
            flag(i, f"年齢範囲外({num})")
            return None
        df["age"] = [clean_age(v, i) for i, v in enumerate(df["age"])]

    if "annual_income" in df.columns:
        def clean_income(s, i):
            if s in ("", "N/A", "n/a", "NA", "null", "-"):
                flag(i, "年収空欄" if s == "" else f"年収非数値({s})")
                return None
            t = unicodedata.normalize("NFKC", s).replace("円", "").replace(",", "").strip()
            try:
                num = int(float(t))
            except ValueError:
                flag(i, f"年収非数値({s})")
                return None
            if num <= 0:
                flag(i, f"年収0以下({num})")
                return None
            return num
        df["annual_income"] = [clean_income(v, i) for i, v in enumerate(df["annual_income"])]

    def parse_date(s, i, col):
        if s in SENTINELS:
            if s != "":
                flag(i, f"{col}:無効値({s})")
            return None
        t = s.strip()
        m = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日", t)
        if m:
            t = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        t = t.replace("/", "-")
        m = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", t)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            t = f"{y}-{mo}-{d}"
        dt = pd.to_datetime(t, format="%Y-%m-%d", errors="coerce")
        if pd.isna(dt):
            dt = pd.to_datetime(t, errors="coerce")
        if pd.isna(dt):
            flag(i, f"{col}:日付解析不可({s})")
            return None
        return dt.strftime("%Y-%m-%d")

    if "registered_date" in df.columns:
        df["registered_date"] = [parse_date(v, i, "登録日") for i, v in enumerate(df["registered_date"])]
    if "last_purchase_date" in df.columns:
        df["last_purchase_date"] = [parse_date(v, i, "最終購入日") for i, v in enumerate(df["last_purchase_date"])]

    if "registered_date" in df.columns and "last_purchase_date" in df.columns:
        for i in range(n):
            r, l = df.at[i, "registered_date"], df.at[i, "last_purchase_date"]
            if isinstance(r, str) and isinstance(l, str) and r > l:
                flag(i, "登録日>最終購入日")

    if "status" in df.columns:
        def clean_status(s, i):
            key = s.strip().lower()
            if key in STATUS_MAP:
                return STATUS_MAP[key]
            if key in ("", "unknown"):
                flag(i, "ステータス不明")
                return "unknown"
            flag(i, f"ステータス不明値({s})")
            return "unknown"
        df["status"] = [clean_status(v, i) for i, v in enumerate(df["status"])]

    if "notes" in df.columns:
        df["notes"] = df["notes"].map(lambda s: re.sub(r"[\r\n]+", " ", s).strip())

    dup_full_mask = df.duplicated(keep="first")
    for i in df.index[dup_full_mask]:
        flag(i, "完全重複行")
    if "customer_id" in df.columns:
        dup_id_mask = df.duplicated(subset=["customer_id"], keep=False) & ~dup_full_mask
        for i in df.index[dup_id_mask]:
            flag(i, "ID重複(要確認)")

    df["品質フラグ"] = ["; ".join(f) for f in flags]
    clean_df = df[~dup_full_mask].reset_index(drop=True)

    def needs_review(flag_list):
        actionable = [f for f in flag_list if not f.startswith("カナがプレースホルダ")]
        return len(actionable) > 0

    review_mask = pd.Series([needs_review(f) for f in flags], index=df.index)
    review_df = df[review_mask].reset_index(drop=True)

    issue_counter = Counter()
    for f in flags:
        for item in f:
            key = re.sub(r"\(.*?\)", "", item)
            issue_counter[key] += 1
    report_rows = sorted(issue_counter.items(), key=lambda x: -x[1])
    report_df = pd.DataFrame(report_rows, columns=["問題の種類", "件数"])

    no_issue = sum(1 for f in flags if len(f) == 0)
    kana_only = sum(1 for f in flags if f and all(x.startswith("カナがプレースホルダ") for x in f))
    summary_df = pd.DataFrame({
        "項目": [
            "入力行数",
            "クレンジング済 行数(完全重複除外後)",
            "完全重複により除外した行数",
            "要確認(個別対応が必要)行数",
            "カナ未整備のみの行数(系統的・個別対応不要)",
            "全くフラグの無い行数",
        ],
        "値": [
            len(raw), len(clean_df), int(dup_full_mask.sum()),
            len(review_df), kana_only, no_issue,
        ],
    })

    return CleanseResult(raw, clean_df, review_df, report_df, summary_df, missing_columns)


def build_excel(result: CleanseResult) -> io.BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    summary_df = result.summary_df
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        result.raw_df.to_excel(writer, sheet_name="元データ", index=False)
        result.clean_df.to_excel(writer, sheet_name="クレンジング済", index=False)
        summary_df.to_excel(writer, sheet_name="レポート", index=False, startrow=0)
        result.report_df.to_excel(writer, sheet_name="レポート", index=False, startrow=len(summary_df) + 3)
        result.review_df.to_excel(writer, sheet_name="要確認", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        sub_font = Font(bold=True, color="1F4E78")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        def style_sheet(ws, header_row=1):
            for cell in ws[header_row]:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = border
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            for col_cells in ws.columns:
                letter = get_column_letter(col_cells[0].column)
                maxlen = 0
                for c in col_cells:
                    if c.value is not None:
                        s = str(c.value)
                        w = sum(2 if unicodedata.east_asian_width(ch) in "WF" else 1 for ch in s)
                        maxlen = max(maxlen, w)
                ws.column_dimensions[letter].width = min(max(maxlen + 2, 8), 45)

        style_sheet(wb["元データ"])
        style_sheet(wb["クレンジング済"])
        style_sheet(wb["要確認"])

        rep = wb["レポート"]
        rep["A1"].fill = header_fill
        rep["A1"].font = header_font
        rep["B1"].fill = header_fill
        rep["B1"].font = header_font
        hdr2 = len(summary_df) + 4
        rep.cell(row=hdr2, column=1).fill = header_fill
        rep.cell(row=hdr2, column=1).font = header_font
        rep.cell(row=hdr2, column=2).fill = header_fill
        rep.cell(row=hdr2, column=2).font = header_font
        rep.cell(row=len(summary_df) + 3, column=1, value="■ 問題種別ごとの件数").font = sub_font
        rep.column_dimensions["A"].width = 40
        rep.column_dimensions["B"].width = 12

        for sheet_name in ("クレンジング済", "要確認"):
            ws = wb[sheet_name]
            flag_col = None
            for cell in ws[1]:
                if cell.value == "品質フラグ":
                    flag_col = cell.column
                    break
            if flag_col:
                warn_fill = PatternFill("solid", fgColor="FFF2CC")
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=flag_col)
                    if c.value:
                        c.fill = warn_fill
    buf.seek(0)
    return buf
